#!/usr/bin/env python3

import openpyxl, os, sys
from pathlib import Path

def main(path_input, path_output):
    sheet_input = openpyxl.load_workbook(path_input).active
    book_output = openpyxl.Workbook()
    os.makedirs(path_output.parent, exist_ok=True)
    for row in range(1, sheet_input.max_row + 1):
        for col in range(1, sheet_input.max_column + 1):
            value = sheet_input.cell(row=row, column=col).value
            book_output.active.cell(row=col, column=row).value = value
    book_output.save(path_output)
    
if __name__ == '__main__':
    main(Path('input/texts.xlsx'), Path('output/swapped_texts.xlsx'))
