#!/usr/bin/env python3

import openpyxl
from pathlib import Path

class Documents():
    def __init__(self, customer, details):
        self.path = Path('output/documents.xlsx')
        self.customer = customer
        self.details = details
        self.issue_documents()

    def issue_documents(self):
        book = openpyxl.load_workbook(self.path)
        for title in ['見積書', '納品書', '請求書']:
            self.issue_document(book[title])
        book.save(self.path)

    def issue_document(self, sheet):
        sheet.cell(row=5, column=2).value = self.customer
        for i, detail in enumerate(self.details, 17):
            for j, value in zip([3, 9, 10, 11], detail):
                sheet.cell(row=i, column=j).value = value
    
if __name__ == '__main__':
    customer = '秋田 太郎'
    details = [('製品A', 10, '台', 1000),
               ('製品B', 20, '個', 2000),
               ('製品C', 30, '式', 3000)]
    d = Documents(customer, details)
